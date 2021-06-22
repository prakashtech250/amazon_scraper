import requests
import os
from sys import platform
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

from scrapy.selector import Selector

USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36'

class amazonApi:
    def __init__(self):
        self.sample_excel = '{}/sample.xlsx'.format(os.getcwd())
        self.filename = None
        self.excel_filename = None
        self.sheet_title = []

    def get_response(self, link):
        headers = {
            'authority': 'www.amazon.com',
            'User-Agent': USER_AGENT,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Cache-Control': 'max-age=0',
        }
        response = requests.get(link, headers=headers)
        res = Selector(text=response.text)
        return res
        
    def create_filename(self):
        if not os.path.exists('output'):
            os.makedirs('output')
        now = datetime.now()
        dt_string = now.strftime("%Y%m%d_%H_%M")
        filename = 'output/{}.xlsx'.format(dt_string)
        return filename

    def clean_data(self, data):
        output = []
        for d in data:
            if d:
                output.append(d.strip().replace('\n',' '))
            else:
                output.append('')
        return output
    
    def get_details(self, asin):
        url = 'https://www.amazon.com/dp/{}'.format(asin)
        response = self.get_response(url)
        title = response.css('#productTitle::text').get()
        if title:
            product_status = response.css('#availability span::text').get()
            rating = response.css('#averageCustomerReviews .a-icon-alt::text').get()
            rating = rating.split(' ')[0]
            detail_bullets = response.css('#detailBullets_feature_div ul li span::text , #detailBullets_feature_div ul li span a::text').extract()
            detail_bullet = self.clean_data(detail_bullets)
            d = 0
            bsr = ''
            for det in detail_bullet:
                if det == 'Best Sellers Rank:':
                    det1 = detail_bullet[d+1].replace('(',' ')
                    det2 = detail_bullet[d+4]
                    det3 = detail_bullet[d+5]
                    bsr = det1 + det2 + ' ' +det3
                    break
                else:
                    d += 1
            price = response.css('#priceblock_ourprice::text').get()
            if price is None:
                price = response.css('#price_inside_buybox::text').get()
            shipping_time = response.css('#mir-layout-DELIVERY_BLOCK-slot-DELIVERY_MESSAGE b::text').get()
            image_url = response.css('.imgTagWrapper img::attr(src)').get()
            data = [product_status, bsr ,rating,price,shipping_time,image_url]
            product_detail = self.clean_data(data)
            return product_detail
        else:
            print('No product')
            return None
    
    def get_manager_email(self, wb):
        asin_list = []
        ws = wb.active
        asin_column = ws['A']
        for x in range(1,len(asin_column)):
            asin = asin_column[x].value
            asin_list.append(asin)
        return asin_list
    
    def get_account_name(self, wb):
        asin_list = []
        ws = wb.active
        asin_column = ws['B']
        for x in range(1,len(asin_column)):
            asin = asin_column[x].value
            asin_list.append(asin)
        return asin_list

    def get_product_group(self, wb):
        asin_list = []
        ws = wb.active
        asin_column = ws['C']
        for x in range(1,len(asin_column)):
            asin = asin_column[x].value
            asin_list.append(asin)
        return asin_list

    def get_asin(self, wb):
        asin_list = []
        ws = wb.active
        asin_column = ws['D']
        for x in range(1,len(asin_column)):
            asin = asin_column[x].value
            asin_list.append(asin)
        return asin_list
    
    def print_banner(self,asin ,data):
        print()
        print('*'*80)
        print('Asin: {}'.format(asin))
        print('Product Status: {}'.format(data[0]))
        print('BSR: {}'.format(data[1]))
        print('Star Rating: {}'.format(data[2]))
        print('Price: {}'.format(data[3]))
        print('Shipping Time: {}'.format(data[4]))
        print('Image Url: {}'.format(data[5]))
        print('*'*80)
                  
    def main(self):
        sample_workbook = load_workbook(self.sample_excel)
        output_filename = self.create_filename()
        wb = Workbook()
        ws = wb.active
        sheet_title = ['Account Manager Email','Account Name','Product Group','Asin to Track','Product Status','Best selling Rating','Star Rating','Price','Shipping Time','Image Url']
        ws.append(sheet_title)
        asin_list = self.get_asin(sample_workbook)
        email_list = self.get_manager_email(sample_workbook)
        account_name = self.get_account_name(sample_workbook)
        product_group = self.get_product_group(sample_workbook)
        count = 0
        for asin in asin_list:
            given_data = [email_list[count],account_name[count],product_group[count],asin]
            data = self.get_details(asin)
            self.print_banner(asin, data)
            ws.append(given_data + data)
            wb.save(output_filename)
            count += 1
        print('Output is saved as {}'.format(output_filename))


if __name__=='__main__':
    api = amazonApi()
    api.main()

    

    

